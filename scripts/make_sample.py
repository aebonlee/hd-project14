# -*- coding: utf-8 -*-
"""신규 딜러 등록 요청 엑셀 샘플 — 일부러 문제 있는 줄을 섞는다."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(BASE, 'sample'); os.makedirs(OUT, exist_ok=True)

wb = Workbook(); ws = wb.active; ws.title = '신규 딜러'
ws['A1'] = '신규 딜러 Access Portal 등록 요청'; ws['A1'].font = Font(bold=True, size=13)
ws['A2'] = '요청일'; ws['B2'] = '2026-08-26'
ws.append([])
head = ['이름', '국가', '언어', '이메일', '딜러사명', '전화번호']
ws.append(head); hr = ws.max_row
fill = PatternFill('solid', fgColor='154B6E'); font = Font(color='FFFFFF', bold=True, size=10)
for c in range(1, len(head)+1):
    cell = ws.cell(row=hr, column=c); cell.fill = fill; cell.font = font
    cell.alignment = Alignment(horizontal='center')

rows = [
    ['Kim Minjun',   'KR', '한국어',  'minjun.kim@daehan.co.kr',   '대한기계',        '+82 10-1111-2222'],
    ['John Miller',  'US', 'English', 'j.miller@midwest-eq.com',   'Midwest Equip.',  '+1 312-555-0134'],
    ['Anna Schmidt', 'DE', 'Deutsch', 'a.schmidt@bau-hd.de',       'Bau HD GmbH',     '+49 89 123456'],
    ['Rahul Sharma', '인도','English', 'rahul@bharat-machines.in',  'Bharat Machines', '+91 22 4000 1234'],
    # 아래는 일부러 문제를 넣었다 (검증이 실제로 잡는지 보려고)
    ['Li Wei',       'ZZ', 'English', 'li.wei@sino-heavy.cn',      'Sino Heavy',      '+86 21 5000 1111'],  # 모르는 국가
    ['Pedro Alves',  'BR', 'Portugu.','pedro alves@br-maq.com.br', 'BR Maquinas',     '+55 11 3000 2222'],  # 이메일 공백
    ['',             'AE', 'English', 'ops@gulf-hd.ae',            'Gulf HD',         '+971 4 123 4567'],   # 이름 없음
    ['Sara Kim',     'KR', '한국어',  'MINJUN.KIM@daehan.co.kr',   '대한기계',        '+82 10-3333-4444'],  # 이메일 중복
    # Region 자동 매핑 시연용 — 정식 템플릿처럼 국가를 영문 전체 이름으로 적으면
    # region-items.json 이 Region 을 찾아 Items to Select 를 제안한다(LA 지역).
    ['Carlos Ruiz',  'Guatemala', 'Spanish', 'c.ruiz@centroam-eq.gt', 'CentroAm Equip.', '+502 2222 3333'],
]
for r in rows: ws.append(r)
for i, w in enumerate([16, 10, 12, 34, 20, 20], start=1):
    ws.column_dimensions[chr(64+i)].width = w
p = os.path.join(OUT, '신규딜러_등록요청_20260826.xlsx')
wb.save(p)
print('  ' + os.path.relpath(p, BASE) + '  ' + str(os.path.getsize(p)) + ' bytes')
